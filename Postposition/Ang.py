import openpyxl
wb = openpyxl.Workbook("FoxNews")
total = wb.create_sheet('total', 1)
Politics = wb.create_sheet('Politics', 2)
Entertainment = wb.create_sheet('Entertainment', 3)
Business = wb.create_sheet('Business', 4)

total.cell(row=1, column=1, value='단어')
total.cell(row=1, column=2, value='출현 횟수')
total.cell(row=1, column=3, value='빈도(%)\n(소수점 2자리까지)')

wb.save('./Politics_FoxNews.xlsx')
